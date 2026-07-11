#!/usr/bin/env python3
"""
generate_ranked_list.py  —  Interactive Pet Marketplace

Generates a filtered "Best for" ranked list spreadsheet.

Usage (run from repo root):
    python Documentation/generate_ranked_list.py "Best for Seniors Living Alone"
    python Documentation/generate_ranked_list.py "Best Budget Friendly Pets"
    python Documentation/generate_ranked_list.py all

The argument is the name of the weighting table file (with or without .xlsx),
or the word "all" to regenerate the Best For list for every weighting table
found in:  Documentation/Best For Weighting Tables/
In "all" mode, the Product Matrix, Product Feature Scores, and Rubric are
loaded once and reused across every table for speed. Each table is processed
independently — if one fails (e.g. a malformed sheet), the rest still run,
and a pass/fail summary prints at the end.

Output spreadsheet saved to:  Documentation/Best for Lists/<name> List.xlsx

Weighting table structure:
    Sheet 1 (any name) — Feature weights: Feature | Suggested Weight
    Sheet "Filters"    — Optional pre-filters: Filter | Operator | Value
                         Operator: equal | not equal | greater than | less than
                         e.g. Price Category | equal | Budget Friendly
                              Rating          | greater than | 4.0
                         Operator column is optional — defaults to 'equal' if blank.
                         If this sheet is absent, no pre-filter is applied.

Filter rule (per Product_Data_Rules.md §4):
    One product per Manufacturer x Price Category x Animal Category.
    Tiebreaker: Score -> Rating -> Visual Contrast -> Review Count.
"""

import sys
import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ── Paths (relative to this script's Documentation/ folder) ───────────────
DOC         = os.path.dirname(os.path.abspath(__file__))
PM_FILE     = os.path.join(DOC, "Product Matrix.xlsx")
SCORES_FILE = os.path.join(DOC, "Product Features Scores", "Product_Feature_Scores.xlsx")
RUBRIC_FILE = os.path.join(DOC, "Feature Scoring Rubric", "Feature_Scoring_Rubric.xlsx")
RULES_FILE  = os.path.join(DOC, "Product_Data_Rules.md")
WT_DIR      = os.path.join(DOC, "Best For Weighting Tables")
OUT_DIR     = os.path.join(DOC, "Best for Lists")

# Feature name aliases: weighting-file name -> scores-file column name (lowercase)
WEIGHT_ALIASES = {
    "behavioral realism level": "realism level",
}

def normalise(name):
    n = name.strip().lower()
    return WEIGHT_ALIASES.get(n, n)

# ── Loaders ────────────────────────────────────────────────────────────────
def resolve_weights_path(table_name):
    name = table_name.strip()
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    path = os.path.join(WT_DIR, name)
    if not os.path.exists(path):
        available = [f for f in os.listdir(WT_DIR) if f.endswith(".xlsx")]
        raise FileNotFoundError(
            f"Weighting table not found: {path}\n"
            f"Available tables:\n" + "\n".join(f"  {f}" for f in available)
        )
    return path, os.path.splitext(table_name.strip())[0]

def load_weights_and_filters(path):
    """
    Returns (weights dict, filters dict).
    Reads the first sheet for weights and the 'Filters' sheet (if present)
    for pre-filters. Filter keys and values are stripped and lowercased for
    comparison, but the original values are kept for display.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Weights (first sheet) ──────────────────────────────────────────────
    ws_w = wb.worksheets[0]
    weights = {}
    for row in ws_w.iter_rows(min_row=2, values_only=True):
        feat, wt = row[0], row[1]
        if feat and wt and isinstance(wt, (int, float)) and wt > 0:
            key = normalise(feat)
            if key != "total":
                weights[key] = float(wt)

    # ── Filters (optional sheet named "Filters") ───────────────────────────
    filters = []   # list of (field, operator, value) tuples
    if "Filters" in wb.sheetnames:
        ws_f = wb["Filters"]
        for row in ws_f.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            field    = str(row[0]).strip()
            # Column B is operator (new), column C is value — or column B is value (legacy 2-col)
            if len(row) >= 3 and row[1] and row[2] is not None:
                operator = str(row[1]).strip().lower()
                value    = str(row[2]).strip()
            else:
                # Legacy 2-column format: default operator to 'equal'
                operator = "equal"
                value    = str(row[1]).strip() if row[1] is not None else ""
            if field and value:
                filters.append((field, operator, value))

    return weights, filters

def load_scores(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[2]]
    rows = {}
    for r in range(3, ws.max_row + 1):
        row  = [ws.cell(row=r, column=c + 1).value for c in range(len(headers))]
        mfr  = str(row[0]).strip().lstrip() if row[0] else ""
        prod = str(row[1]).strip()          if row[1] else ""
        if mfr or prod:
            rows[(mfr.lower(), prod.lower())] = {
                headers[i]: row[i] for i in range(len(headers))
            }
    return headers, rows

def load_product_matrix(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Product Matrix"]
    h = [c.value for c in ws[1]]
    pm = {}
    for r in range(2, ws.max_row + 1):
        row  = [ws.cell(row=r, column=c + 1).value for c in range(len(h))]
        mfr  = str(row[1]).strip() if row[1] else ""
        prod = str(row[2]).strip() if row[2] else ""
        if mfr or prod:
            # Normalised aliases kept for backward compatibility
            entry = {
                "category":  row[3],   # Animal Category
                "type":      row[4],   # Product Type (Fluffy Companion / Ai & Robotic Pets)
                "price_cat": row[20],  # Price Category
                "rating":    row[12],  # Rating
                "reviews":   row[13],  # Review Count
                "price":     row[17],  # Price
            }
            # All columns by their exact header name so any PM field works as a filter
            for i, header in enumerate(h):
                if header:
                    entry[str(header).strip()] = row[i]
            pm[(mfr.lower(), prod.lower())] = entry
    return pm

def load_rubric(path):
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rubric = {}
    for r in range(4, ws.max_row + 1):
        feat  = ws.cell(row=r, column=1).value
        desc  = ws.cell(row=r, column=2).value
        stype = ws.cell(row=r, column=3).value
        if feat:
            rubric[normalise(feat)] = {
                "description": desc  or "",
                "score_type":  stype or "1-5",
            }
    return rubric

# ── Pre-filter ─────────────────────────────────────────────────────────────
# Supported field names (case-insensitive) → pm_data key
FILTER_FIELD_MAP = {
    "price category":  "price_cat",
    "animal category": "category",
    "category":        "category",
    "rating":          "rating",
    "reviews":         "reviews",
    "price":           "price",
}

def _compare(actual, operator, target):
    """
    Compare actual vs target using the given operator.
    Tries numeric comparison first; falls back to case-insensitive string.
    Operators: equal | not equal | greater than | less than | contains |
               is blank | is not blank
    """
    op = operator.strip().lower()
    # Blank / not-blank — check before any string conversion
    if op in ("is blank", "blank", "empty"):
        return actual is None or str(actual).strip().lower() in ("", "none", "nan")
    if op in ("is not blank", "not blank", "not empty"):
        return actual is not None and str(actual).strip().lower() not in ("", "none", "nan")
    # Contains
    if op == "contains":
        return target.strip().lower() in str(actual).strip().lower()
    # Attempt numeric
    try:
        a = float(str(actual).replace("$","").replace(",","").strip())
        t = float(str(target).replace("$","").replace(",","").strip())
        if op == "equal":        return a == t
        if op == "not equal":    return a != t
        if op == "greater than": return a >  t
        if op == "less than":    return a <  t
    except (ValueError, TypeError):
        pass
    # String comparison (case-insensitive)
    a = str(actual).strip().lower()
    t = str(target).strip().lower()
    if op == "equal":        return a == t
    if op == "not equal":    return a != t
    if op == "greater than": return a > t
    if op == "less than":    return a < t
    print(f"  WARNING: Unknown filter operator '{operator}' — row passes by default.")
    return True

def passes_filters(pm_data, filters):
    """
    Returns True if the product passes all filter rules.
    filters is a list of (field, operator, value) tuples.

    Lookup order for each field:
    1. FILTER_FIELD_MAP normalised alias (backward compat)
    2. Exact match against PM column header stored in pm_data
    3. Case-insensitive match against PM column headers
    If the field genuinely cannot be found, a warning is printed and the
    filter is skipped (pass-through) rather than silently ignored.
    """
    for field, operator, value in filters:
        # 1. Normalised alias
        pm_key = FILTER_FIELD_MAP.get(field.lower())
        if pm_key is not None:
            actual = pm_data.get(pm_key)
        # 2. Exact header name
        elif field in pm_data:
            actual = pm_data[field]
        # 3. Case-insensitive header name
        else:
            match = next((v for k, v in pm_data.items() if k.lower() == field.lower()), None)
            if match is not None:
                actual = match
            else:
                print(f"  WARNING: Filter field '{field}' not found in Product Matrix columns — filter skipped.")
                continue
        if not _compare(actual, operator, value):
            return False
    return True

# ── Scoring ────────────────────────────────────────────────────────────────
def calc_score(score_row, feat_headers, weights):
    total = 0.0
    for feat_norm, wt in weights.items():
        for h in feat_headers:
            if h and normalise(h) == feat_norm:
                val = score_row.get(h)
                if isinstance(val, (int, float)):
                    total += val * wt
                break
    return round(total, 4)

# ── Manufacturer x Price x Animal filter ──────────────────────────────────
def apply_filter(candidates):
    groups = {}
    for c in candidates:
        key = (c["mfr_clean"], c["price_cat"], c["animal_cat"])
        groups.setdefault(key, []).append(c)

    winners, eliminated = [], []
    for key, group in groups.items():
        group.sort(key=lambda x: (
            x["score"],
            x["rating"]       or 0,
            x["vis_contrast"] or 0,
            x["reviews"]      or 0,
        ), reverse=True)
        winners.append(group[0])
        for g in group[1:]:
            g["eliminated_by"] = group[0]["prod_clean"]
            eliminated.append(g)
    return winners, eliminated

# ── Output spreadsheet ─────────────────────────────────────────────────────
def write_output(winners, eliminated, pre_filtered, weights, filters,
                 rubric, list_name, out_path):
    wb = openpyxl.Workbook()

    hdr_font   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    meta_font  = Font(name="Calibri", size=10, italic=True, color="595959")
    body_font  = Font(name="Calibri", size=10)
    bold_font  = Font(name="Calibri", size=10, bold=True, color="1F4E79")
    elim_font  = Font(name="Calibri", size=10, color="999999")
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    alt_fill   = PatternFill("solid", fgColor="EBF3FB")
    wt_fill    = PatternFill("solid", fgColor="E2EFDA")
    flt_fill   = PatternFill("solid", fgColor="FFF2CC")
    elim_fill  = PatternFill("solid", fgColor="F5F5F5")
    wrap       = Alignment(wrap_text=True, vertical="top")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin       = Side(style="thin", color="BFBFBF")
    bdr        = Border(top=thin, bottom=thin, left=thin, right=thin)

    feat_cols = [(fn, rubric.get(fn, {}).get("description", fn)[:28], 10)
                 for fn in weights]

    # ── Sheet 1: Ranked List ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Ranked List"
    ws.sheet_view.showGridLines = False

    # Title
    total_cols = 9 + len(feat_cols) + 1
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"{list_name} — Ranked Product List"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    # Meta row
    filter_note = (f"  |  Pre-filter: {', '.join(f'{f} {op} {v}' for f,op,v in filters)}"
                   f"  ({pre_filtered} products removed)" if filters else "")
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    ws["A2"] = (f"Generated: {date.today().isoformat()}"
                f"  |  Products shown: {len(winners)}"
                f"  |  Eliminated by dedup rule: {len(eliminated)}"
                f"{filter_note}")
    ws["A2"].font = meta_font
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 20

    # Pre-filter row (only if filters were applied)
    data_start_row = 3
    if filters:
        ws.merge_cells(f"A3:{get_column_letter(total_cols)}3")
        filter_str = "  |  ".join(f"{f} {op} {v}" for f, op, v in filters)
        ws["A3"] = f"Pre-filter applied (before scoring): {filter_str}"
        ws["A3"].font = Font(name="Calibri", size=10, bold=True, color="7B6000")
        ws["A3"].fill = flt_fill
        ws["A3"].alignment = Alignment(vertical="center")
        ws.row_dimensions[3].height = 18
        data_start_row = 4

    # Weights row
    ws.merge_cells(f"A{data_start_row}:{get_column_letter(min(3, total_cols))}{data_start_row}")
    ws[f"A{data_start_row}"] = "Weights:"
    ws[f"A{data_start_row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"A{data_start_row}"].fill = wt_fill
    ws[f"A{data_start_row}"].alignment = Alignment(vertical="center")
    col = 4
    for fn, wt in weights.items():
        c = ws.cell(row=data_start_row, column=col,
                    value=f"{fn}: {wt:.0%}")
        c.font = Font(name="Calibri", size=10)
        c.fill = wt_fill
        col += 1
    ws.row_dimensions[data_start_row].height = 16

    # Column headers
    hdr_row = data_start_row + 1
    fixed_cols = [
        ("Rank", 5), ("Manufacturer", 20), ("Product", 32),
        ("Price\nCategory", 12), ("Category", 10), ("Price", 9),
        ("Rating", 8), ("Reviews", 9), ("Overall\nScore", 10),
    ]
    notes_col_def = [("Notes / Flags", 40)]
    all_col_defs  = fixed_cols + [(fc[0], fc[2]) for fc in feat_cols] + notes_col_def

    for ci, (hdr, width) in enumerate(all_col_defs, 1):
        c = ws.cell(row=hdr_row, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[hdr_row].height = 30
    ws.freeze_panes = f"A{hdr_row + 1}"

    # Data rows
    for ri, w in enumerate(winners, hdr_row + 1):
        bg = alt_fill if ri % 2 == 0 else None

        def set_cell(col, val, fnt=body_font, aln=None):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = fnt; c.border = bdr
            c.alignment = aln or wrap
            if bg: c.fill = bg

        set_cell(1, ri - hdr_row, bold_font, center)
        set_cell(2, w["mfr_clean"])
        set_cell(3, w["prod_clean"])
        set_cell(4, w["price_cat"],  aln=center)
        set_cell(5, w["animal_cat"], aln=center)
        set_cell(6, w["price"],      aln=center)
        set_cell(7, w["rating"],     aln=center)
        set_cell(8, w["reviews"],    aln=center)

        sc = ws.cell(row=ri, column=9, value=round(w["score"], 2))
        sc.font = Font(name="Calibri", size=11, bold=True,
                       color="1D7044" if w["score"] >= 4.0 else
                             "185FA5" if w["score"] >= 3.0 else "595959")
        sc.border = bdr; sc.alignment = center
        if bg: sc.fill = bg

        for fi, (fn, _, _) in enumerate(feat_cols, 10):
            c = ws.cell(row=ri, column=fi,
                        value=w.get("feat_scores", {}).get(fn))
            c.font = body_font; c.border = bdr; c.alignment = center
            if bg: c.fill = bg

        notes_ci = 10 + len(feat_cols)
        set_cell(notes_ci, w.get("notes", ""))
        ws.row_dimensions[ri].height = 45

    # ── Sheet 2: Eliminated ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Eliminated")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:H1")
    ws2["A1"] = (f"Eliminated by dedup rule — {list_name}  "
                 "(same Manufacturer x Price Category x Animal Category "
                 "as a higher-scoring product)")
    ws2["A1"].font = Font(name="Calibri", size=11, bold=True, color="999999")
    ws2["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[1].height = 22

    e_hdrs = [("Manufacturer",18),("Product",32),("Price Category",14),
              ("Category",10),("Score",9),("Rating",8),("Kept Instead",32),("Reason",30)]
    for ci, (h, w) in enumerate(e_hdrs, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = bdr
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 22

    for ri, e in enumerate(sorted(eliminated, key=lambda x: x["score"], reverse=True), 3):
        vals = [e["mfr_clean"], e["prod_clean"], e["price_cat"], e["animal_cat"],
                round(e["score"], 2), e["rating"], e.get("eliminated_by", ""),
                f"score={e['score']:.2f} rating={e['rating']}"]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = elim_font; c.border = bdr
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if ri % 2 == 0: c.fill = elim_fill
        ws2.row_dimensions[ri].height = 18

    # ── Sheet 3: Weights Used ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Weights Used")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"Weights applied — {list_name}"
    ws3["A1"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    ws3["A1"].alignment = Alignment(vertical="center")
    ws3.row_dimensions[1].height = 20

    if filters:
        ws3.merge_cells("A2:D2")
        ws3["A2"] = "Pre-filter: " + "  |  ".join(f"{f} {op} {v}" for f, op, v in filters)
        ws3["A2"].font = Font(name="Calibri", size=10, bold=True, color="7B6000")
        ws3["A2"].fill = flt_fill
        ws3["A2"].alignment = Alignment(vertical="center")
        ws3.row_dimensions[2].height = 16
        wt_hdr_row = 3
    else:
        wt_hdr_row = 2

    for ci, (h, w) in enumerate([("Feature (scores-file column)",32),
                                   ("Weight",10),("What It Measures",50),
                                   ("Score Type",14)], 1):
        c = ws3.cell(row=wt_hdr_row, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[wt_hdr_row].height = 18

    total_wt = 0.0
    for ri, (fn, wt) in enumerate(weights.items(), wt_hdr_row + 1):
        rb = rubric.get(fn, {})
        for ci, val in enumerate([fn, f"{wt:.0%}",
                                   rb.get("description",""),
                                   rb.get("score_type","")], 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = body_font
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ri % 2 == 0: c.fill = alt_fill
        ws3.row_dimensions[ri].height = 20
        total_wt += wt

    tr = wt_hdr_row + 1 + len(weights)
    ws3.cell(row=tr, column=1, value="TOTAL").font = Font(name="Calibri", size=10, bold=True)
    ws3.cell(row=tr, column=2, value=f"{total_wt:.0%}").font = Font(name="Calibri", size=10, bold=True)

    os.makedirs(OUT_DIR, exist_ok=True)
    try:
        wb.save(out_path)
    except PermissionError:
        print()
        print('ERROR: Cannot save output file — it is open in Excel.')
        print(f'  Close this file in Excel and run the script again:')
        print(f'  {out_path}')
        raise SystemExit(1)

# ── TypeScript code generation ─────────────────────────────────────────────
# When generate_one() runs for a list name that has a TS config registered
# here, it also regenerates the corresponding TypeScript data files so that
# the Next.js site always stays in sync with the spreadsheet data.
#
# To add a new Best For page: copy the "Best for Seniors in Memory Care
# Facilities" entry below, update the paths and criteria_display, and the
# script will regenerate its data files on every run.

import re as _re

REPO_ROOT = os.path.dirname(DOC)  # one level up from Documentation/


SITE_URL = "https://interactivepetmarketplace.com"

# Keywords that indicate a feature is "reversed" — score 5 = lowest risk/burden,
# not highest amount of that property.
_REVERSED_KEYWORDS = {"risk", "burden", "maintenance", "difficulty", "fall"}

def _to_slug(name):
    """Match generate-products.js toSlug() exactly — keep in sync."""
    s = str(name).strip().lower().replace("&", "and")
    s = _re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")

def _is_reversed(feature_name):
    """Heuristic: features containing risk/burden/maintenance keywords are reversed."""
    return bool(set(feature_name.lower().split()) & _REVERSED_KEYWORDS)

def _auto_key(feature_name):
    """Generate a short TypeScript-safe identifier from a normalised feature name.
    'caregiver burden' → 'caregiver'   'safety risk' → 'safety'
    'charging convenience' → 'charging'  'cleanability' → 'cleanability'
    """
    words = _re.sub(r"[^a-z0-9 ]+", "", feature_name.strip().lower()).split()
    stop = {"the","a","an","of","for","and","or","in","with","level","type"}
    sig = [w for w in words if w not in stop] or words
    return sig[0][:12]  # first significant word, max 12 chars

def _auto_label(feature_name):
    """Title-case label from normalised name. 'caregiver burden' → 'Caregiver Burden'"""
    return feature_name.strip().title()

def _derive_criteria(weights):
    """Auto-derive criteria_display from the weights dict (no manual config needed)."""
    seen_keys = {}
    result = {}
    for fn_norm in weights:
        key = _auto_key(fn_norm)
        # Deduplicate keys if two features produce the same short key
        if key in seen_keys:
            key = _re.sub(r"[^a-z0-9]", "", fn_norm)[:14]
        seen_keys[key] = fn_norm
        label = _auto_label(fn_norm)
        result[fn_norm] = {
            "key":      key,
            "label":    label,
            "short":    label,
            "reversed": _is_reversed(fn_norm),
        }
    return result

def _load_existing_slugs():
    """Return the set of slugs in components/site-data.ts, or None if absent."""
    path = os.path.join(REPO_ROOT, "components", "site-data.ts")
    if not os.path.exists(path):
        print("  WARNING: site-data.ts not found — slug validation skipped.")
        return None
    with open(path, encoding="utf-8") as f:
        return set(_re.findall(r'"slug":\s*"([^"]+)"', f.read()))

def _date_display():
    today = date.today()
    months = ["January","February","March","April","May","June","July",
              "August","September","October","November","December"]
    return f"{months[today.month - 1]} {today.day}, {today.year}", today.isoformat()

def _get_score(srow, feat_hdrs, substr):
    """Return an integer score from srow by matching a column name substring."""
    for h in feat_hdrs:
        if h and substr.lower() in str(h).lower():
            v = srow.get(h)
            return int(v) if isinstance(v, (int, float)) else 0
    return 0

def _find_existing_page_dir(list_name):
    """
    Scan app/ for a Best For page directory that was previously generated for
    this list. Detection: looks for page-data.ts containing the script marker.
    Returns the directory path if found, None otherwise.
    """
    app_dir = os.path.join(REPO_ROOT, "app")
    if not os.path.isdir(app_dir):
        return None
    marker = f'generate_ranked_list.py "{list_name}"'
    for entry in sorted(os.listdir(app_dir)):
        pd = os.path.join(app_dir, entry, "page-data.ts")
        if os.path.isfile(pd):
            with open(pd, encoding="utf-8") as f:
                if marker in f.read():
                    return os.path.join(app_dir, entry)
    return None

def _get_page_dir(list_name):
    """Return (page_dir, page_slug, is_new_page)."""
    existing = _find_existing_page_dir(list_name)
    if existing:
        return existing, os.path.basename(existing), False
    slug = _to_slug(list_name)
    return os.path.join(REPO_ROOT, "app", slug), slug, True

def _build_json_ld(list_name, page_slug, ranked, top_score):
    """Build the JSON-LD structured data string for this Best For page."""
    import json
    page_url = f"{SITE_URL}/{page_slug}"
    items = []
    for i, (slug, pct, w, p) in enumerate(ranked):
        item = {
            "@type": "ListItem",
            "position": i + 1,
            "item": {
                "@type": "Product",
                "name": w["prod_clean"],
                "brand": {"@type": "Brand", "name": w["mfr_clean"]},
                "offers": {
                    "@type": "Offer",
                    "price": f"{p['price']:.2f}",
                    "priceCurrency": "USD",
                    "availability": "https://schema.org/InStock",
                },
            },
        }
        if p.get("rating") and float(p["rating"]) > 0:
            item["item"]["aggregateRating"] = {
                "@type": "AggregateRating",
                "ratingValue": f"{float(p['rating']):.1f}",
                "reviewCount": str(p["reviews"]),
                "bestRating": "5",
                "worstRating": "1",
            }
        items.append(item)
    schema = {
        "@context": "https://schema.org",
        "@type": "ItemList",
        "name": list_name,
        "url": page_url,
        "numberOfItems": len(items),
        "itemListElement": items,
    }
    return json.dumps(schema, separators=(",", ":"))

def _write_page_template(page_dir, page_slug, list_name, crit, accent="purple"):
    """Create page.tsx and scoring/page.tsx templates for a brand-new Best For page."""
    page_path = os.path.join(page_dir, "page.tsx")
    if os.path.exists(page_path):
        return  # Never overwrite an existing hand-crafted page

    os.makedirs(os.path.join(page_dir, "scoring"), exist_ok=True)

    # Scoring page — imports from ./scoring-data, identical structure to memory-care
    scoring_tpl = (
        'import { PageShell } from "@/components/layout";\n'
        'import Link from "next/link";\n'
        '// AUTO-GENERATED data\n'
        'import { WEIGHTS, ROWS, GENERATED_DATE, type ScoreKey } from "./scoring-data";\n\n'
        'function ScoreCell({ value, reversed }: { value: number; reversed: boolean }) {\n'
        '  const isTop = value === 5;\n'
        '  const isLow = value <= 2;\n'
        '  return (\n'
        '    <td className={`border border-trust-200 px-3 py-2 text-center text-sm font-semibold ${\n'
        '      isTop ? "bg-trust-50 text-trust-700" : isLow ? "text-slate-400" : "text-slate-700"\n'
        '    }`}>{value}</td>\n'
        '  );\n'
        '}\n\n'
        'export default function BestForScoring() {\n'
        '  return (\n'
        '    <PageShell>\n'
        '      <section className="section-pad">\n'
        '        <div className="container-shell">\n'
        f'          <div className="mb-8"><Link href="/{page_slug}" className="text-sm font-semibold text-trust-600 hover:underline">← Back to rankings</Link></div>\n'
        '          <p className="eyebrow">Detailed Scoring</p>\n'
        f'          <h1 className="mt-2 text-2xl font-bold tracking-tight text-slate-900 sm:text-3xl">{list_name} — Top 5 Results</h1>\n'
        '          <p className="mt-3 max-w-3xl text-sm leading-7 text-slate-700">\n'
        '            All scores are 1–5. Criteria marked <span className="font-semibold">(↑ higher = better outcome)</span> show 5 as lowest risk/burden.\n'
        '            Generated {GENERATED_DATE}. Tiebreaker: Score → Rating → Visual Contrast → Review Count.\n'
        '          </p>\n'
        '          <div className="mt-8 overflow-x-auto rounded-2xl border border-trust-200 shadow-soft">\n'
        '            <table className="w-full border-collapse text-sm">\n'
        '              <thead><tr className="bg-trust-50">\n'
        '                <th className="border border-trust-200 px-3 py-3 text-left font-semibold text-slate-900">Rank</th>\n'
        '                <th className="border border-trust-200 px-3 py-3 text-left font-semibold text-slate-900">Product</th>\n'
        '                <th className="border border-trust-200 px-3 py-3 text-right font-semibold text-slate-900">Price</th>\n'
        '                <th className="border border-trust-200 px-3 py-3 text-right font-semibold text-slate-900">Rating</th>\n'
        '                <th className="border border-trust-200 px-3 py-3 text-center font-bold text-trust-900">Overall</th>\n'
        '                {WEIGHTS.map((w) => (\n'
        '                  <th key={w.key} className="border border-trust-200 px-3 py-3 text-center font-semibold text-slate-700 whitespace-nowrap">\n'
        '                    {w.label}<br/><span className="text-xs font-normal text-slate-500">{w.weight}</span>\n'
        '                    {w.reversed && <><br/><span className="text-[10px] font-normal text-trust-600">↑ higher = better</span></>}\n'
        '                  </th>\n'
        '                ))}\n'
        '              </tr></thead>\n'
        '              <tbody>\n'
        '                {ROWS.map((row, i) => (\n'
        '                  <tr key={row.rank} className={i % 2 === 0 ? "bg-white" : "bg-cream-50"}>\n'
        '                    <td className="border border-trust-200 px-3 py-3 font-bold text-trust-700">#{row.rank}</td>\n'
        '                    <td className="border border-trust-200 px-3 py-3"><p className="font-semibold text-slate-900">{row.product}</p><p className="text-xs text-slate-500">{row.manufacturer} · {row.animal} · {row.priceCategory}</p></td>\n'
        '                    <td className="border border-trust-200 px-3 py-3 text-right text-slate-700">{row.price}</td>\n'
        '                    <td className="border border-trust-200 px-3 py-3 text-right text-slate-700">★ {row.rating}<br/><span className="text-xs text-slate-500">({row.reviews} reviews)</span></td>\n'
        '                    <td className="border border-trust-200 px-3 py-3 text-center font-bold text-trust-900">{row.overall}</td>\n'
        '                    {WEIGHTS.map((w) => <ScoreCell key={w.key} value={row.scores[w.key]} reversed={w.reversed} />)}\n'
        '                  </tr>\n'
        '                ))}\n'
        '              </tbody>\n'
        '            </table>\n'
        '          </div>\n'
        '          <p className="mt-6 text-xs text-slate-500">Full scoring methodology: Documentation/Feature_Scoring_Rubric.xlsx. Product data: Documentation/Product Matrix.xlsx.</p>\n'
        '        </div>\n'
        '      </section>\n'
        '    </PageShell>\n'
        '  );\n'
        '}\n'
    )
    with open(os.path.join(page_dir, "scoring", "page.tsx"), "w", encoding="utf-8") as f:
        f.write(scoring_tpl)

    # Main page — minimal but functional template
    page_tpl = (
        'import { PageShell } from "@/components/layout";\n'
        'import { BestForCard } from "@/components/best-for-card";\n'
        'import { ScoreGauge, ScoreBar } from "@/components/score-gauge";\n'
        'import {\n'
        '  IconSparkleClean, IconDurability, IconUsers, IconShieldCheck,\n'
        '  IconBattery, IconPrivacy, IconBrain, IconHeartHands, IconAward,\n'
        '  featureIcon, StarRating,\n'
        '} from "@/components/best-for-icons";\n'
        'import { products } from "@/components/site-data";\n'
        'import Link from "next/link";\n\n'
        '// ── Auto-generated data (regenerated by generate_ranked_list.py) ─────────\n'
        'import {\n'
        '  SPREADSHEET_UPDATED, TOP_SCORE_IN_GROUP, TOP_PICK_RAW_SCORE, TOP_PICK_PERCENT,\n'
        '  RANKED_SLUGS, SCORE_PERCENT, RUNNER_NOTES, TOP_PICK_CRITERIA_DATA, JSON_LD,\n'
        '} from "./page-data";\n'
        'import { SCORES } from "./scores";\n\n'
        '// ── Icon mapping — update labels to match TOP_PICK_CRITERIA_DATA in page-data.ts ─\n'
        'const ICON_MAP: Record<string, React.ComponentType<{ className?: string }>> = {\n'
        '  "Cleanability": IconSparkleClean, "Durability": IconDurability,\n'
        '  "Caregiver Burden": IconUsers,    "Safety Risk": IconShieldCheck,\n'
        '  "Charging Convenience": IconBattery, "Privacy Risk": IconPrivacy,\n'
        '  "Dementia Suitability": IconBrain,\n'
        '  // TODO: add icons for any criteria specific to this list\n'
        '};\n'
        'const TOP_PICK_CRITERIA = TOP_PICK_CRITERIA_DATA.map((c) => ({\n'
        '  ...c, Icon: ICON_MAP[c.label] ?? IconSparkleClean,\n'
        '}));\n\n'
        '// ── TODO: Review these bullets when the top pick changes ─────────────────\n'
        '// Must be grounded in locked rubric anchors — never invented marketing claims.\n'
        'const WHY_TOP_PICK_BULLETS: { label: string; note: string }[] = [\n'
        '  // TODO: populate from top pick\'s highest-scoring criteria\n'
        '];\n\n'
        'export default function BestForPage() {\n'
        '  const ranked = RANKED_SLUGS\n'
        '    .map((slug) => products.find((p) => p.slug === slug))\n'
        '    .filter(Boolean) as NonNullable<(typeof products)[number]>[];\n'
        '  const topPick = ranked[0];\n'
        '  const runners = ranked.slice(1);\n\n'
        '  return (\n'
        '    <PageShell>\n'
        '      <script type="application/ld+json" dangerouslySetInnerHTML={{ __html: JSON_LD }} />\n\n'
        f'      {{/* Hero */}}\n'
        f'      <section className="bg-{accent}-100 py-8 sm:py-10">\n'
        '        <div className="container-shell">\n'
        '          <div className="flex items-start gap-4">\n'
        f'            <div className="hidden sm:flex h-20 w-20 shrink-0 items-center justify-center rounded-full bg-white text-{accent}-600 shadow-soft">\n'
        '              <IconHeartHands className="h-10 w-10" />\n'
        '            </div>\n'
        '            <div>\n'
        '              <p className="eyebrow">Best For Rankings</p>\n'
        f'              <h1 className="mt-1 text-4xl font-extrabold tracking-tight text-slate-900 sm:text-5xl">{list_name}</h1>\n'
        '              <p className="mt-2 max-w-2xl text-base font-medium leading-7 text-slate-700">\n'
        '                {/* TODO: add description */}\n'
        '              </p>\n'
        '              <p className="mt-2 inline-flex items-center gap-2 text-xs font-semibold text-trust-700">\n'
        '                <span className="inline-flex h-6 w-6 items-center justify-center rounded-full bg-trust-500 text-white text-[10px]">✓</span>\n'
        '                Ratings from Verified Sources · Updated {SPREADSHEET_UPDATED}\n'
        '              </p>\n'
        '            </div>\n'
        '          </div>\n'
        '        </div>\n'
        '      </section>\n\n'
        '      {/* See app/best-for-memory-care/page.tsx for the full Best For page template.\n'
        '          TODO: add Top Pick card, Why Top Pick, How We Ranked, Customize, Also Ranked sections. */}\n\n'
        '      {/* Also Ranked */}\n'
        '      <section className="py-4 sm:py-5 bg-white">\n'
        '        <div className="container-shell">\n'
        '          <h2 className="text-xl font-extrabold tracking-tight text-slate-900 sm:text-2xl">Top Picks</h2>\n'
        '          <p className="mt-0.5 text-sm text-slate-700">Ranked by weighted criteria for this audience.</p>\n'
        '          <div className="mt-2 grid gap-2.5 sm:grid-cols-2 xl:grid-cols-5">\n'
        '            {ranked.map((product, i) => (\n'
        '              <div key={product.slug} className="flex flex-col">\n'
        '                <p className="mb-1 text-xs font-bold uppercase tracking-wide text-slate-500">#{i + 1} Ranked</p>\n'
        f'                <BestForCard product={{product}} note={{RUNNER_NOTES[product.slug]}} className="flex-1"\n'
        f'                             scorePercent={{SCORE_PERCENT[product.slug]}} accentColor="text-{accent}-600" />\n'
        '              </div>\n'
        '            ))}\n'
        '          </div>\n'
        f'          <div className="mt-4 border-t border-{accent}-200 pt-3 text-center">\n'
        f'            <Link href="/{page_slug}/scoring" className="text-sm font-semibold text-trust-600 underline underline-offset-4 hover:text-trust-800">Show detailed scoring results →</Link>\n'
        '          </div>\n'
        '        </div>\n'
        '      </section>\n'
        '    </PageShell>\n'
        '  );\n'
        '}\n'
    )
    with open(page_path, "w", encoding="utf-8") as f:
        f.write(page_tpl)

    print(f"  ✓ (new template) {os.path.relpath(page_path, REPO_ROOT)}")
    print(f"  ✓ (new template) {os.path.relpath(os.path.join(page_dir, 'scoring', 'page.tsx'), REPO_ROOT)}")


def generate_ts_files(list_name, winners, candidates, weights, feat_hdrs, srows, pm):
    """
    Regenerate TypeScript data files for a Best For page.
    Called automatically from generate_one() for every weighting table.

    Auto-discovers the existing page directory by scanning app/ for page-data.ts
    files containing the list name. If no existing page is found, creates a new
    directory at the canonical slug path (slugified list name) with template files.
    """
    print(f"\n  Generating TypeScript data files …")
    date_display, date_iso = _date_display()
    existing_slugs = _load_existing_slugs()
    page_dir, page_slug, is_new = _get_page_dir(list_name)

    # ── Build per-product dataset (all evaluated, not just winners) ──────────
    all_prods, skipped = [], []
    for p in candidates:
        slug = _to_slug(p["prod_clean"])
        if existing_slugs is not None and slug not in existing_slugs:
            skipped.append((p["prod_clean"], slug))
            continue
        mfr_l  = p["mfr_clean"].lower().lstrip()
        prod_l = p["prod_clean"].lower()
        srow   = srows.get((mfr_l, prod_l), {})
        pm_d   = pm.get((mfr_l, prod_l), {})
        all_prods.append({
            "slug":      slug,
            "score":     p["score"],
            "price":     float(pm_d.get("price") or 0),
            "price_cat": p.get("price_cat") or pm_d.get("price_cat") or "Unknown",
            "animal":    p.get("animal_cat") or pm_d.get("category") or "Unknown",
            "type":      pm_d.get("type") or "Unknown",
            "movement":  _get_score(srow, feat_hdrs, "movement level"),
            "sound":     _get_score(srow, feat_hdrs, "sound quality"),
            "visual":    p.get("vis_contrast") or _get_score(srow, feat_hdrs, "visual contrast"),
            "rating":    float(pm_d.get("rating") or 0),
            "reviews":   int(pm_d.get("reviews") or 0),
            "srow":      srow,
        })

    if skipped:
        print(f"  WARNING: {len(skipped)} product(s) not in site-data.ts — excluded:")
        for name, sl in skipped:
            print(f"    {name!r}  →  {sl!r}")
        print("  Run npm run generate:products after updating Product Matrix.")

    if not all_prods:
        print("  SKIP: no products match site-data.ts — TS files not updated.")
        return

    top_score = max(p["score"] for p in all_prods)
    for p in all_prods:
        p["pct"] = round(p["score"] / top_score * 100)

    # Deduped top 5 (winners that exist in site-data.ts)
    ranked = []
    for w in winners:
        sl = _to_slug(w["prod_clean"])
        hit = next((p for p in all_prods if p["slug"] == sl), None)
        if hit:
            ranked.append((sl, hit["pct"], w, hit))
        if len(ranked) == 5:
            break

    if not ranked:
        print("  SKIP: no ranked products found in site-data.ts.")
        return

    top_pick_slug, top_pick_pct, top_pick_w, top_pick_p = ranked[0]

    # Auto-derive criteria display from weights
    crit = _derive_criteria(weights)

    # ── scores.ts (in page directory) ─────────────────────────────────────
    os.makedirs(page_dir, exist_ok=True)
    pc_vals = sorted({p["price_cat"] for p in all_prods} - {"Unknown"})
    ac_vals = sorted({p["animal"]    for p in all_prods} - {"Unknown"})
    ty_vals = sorted({p["type"]      for p in all_prods} - {"Unknown"})
    rows_ts = "\n".join(
        f'  {{ slug: {p["slug"]!r}, score: {p["score"]:.2f}, scorePercent: {p["pct"]},'
        f' price: {p["price"]:.2f}, priceCategory: {p["price_cat"]!r},'
        f' animalCategory: {p["animal"]!r}, type: {p["type"]!r},'
        f' movementLevel: {p["movement"]}, soundQuality: {p["sound"]},'
        f' visualContrast: {p["visual"]} }},'
        for p in all_prods
    )
    pc_u = " | ".join(f'"{v}"' for v in pc_vals)
    ac_u = " | ".join(f'"{v}"' for v in ac_vals)
    ty_u = " | ".join(f'"{v}"' for v in ty_vals)
    scores_content = (
        f'// AUTO-GENERATED — do not edit by hand.\n'
        f'// Run: python Documentation/generate_ranked_list.py "{list_name}"\n'
        f'// Generated: {date_iso}\n\n'
        f'export const TOP_SCORE_IN_GROUP = {top_score};\n\n'
        f'export type BestForScoreRow = {{\n'
        f'  slug: string; score: number; scorePercent: number; price: number;\n'
        f'  priceCategory: {pc_u}; animalCategory: {ac_u}; type: {ty_u};\n'
        f'  movementLevel: number; soundQuality: number; visualContrast: number;\n}};\n\n'
        f'// Alias used by CustomizeRankings\n'
        f'// Legacy alias\nexport type MemoryCareScoreRow = BestForScoreRow;\n'
        f'export const SCORES: BestForScoreRow[] = [\n{rows_ts}\n];\n'
        f'// Legacy alias for existing imports\nexport const MEMORY_CARE_SCORES = SCORES;\n'
    )
    scores_path = os.path.join(page_dir, "scores.ts")
    with open(scores_path, "w", encoding="utf-8") as f:
        f.write(scores_content)
    print(f"  ✓  {os.path.relpath(scores_path, REPO_ROOT)}")

    # ── page-data.ts ──────────────────────────────────────────────────────
    ranked_lines = "\n".join(f'  {s!r},  // Rank {i+1} — {pct}%'
                             for i,(s,pct,_,__) in enumerate(ranked))
    pct_lines    = "\n".join(f'  {s!r}: {pct},' for s,pct,_,__ in ranked)
    crit_lines   = "\n".join(
        f'  {{ label: {crit[fn]["label"]!r}, weight: "{wt:.0%}", score: {_get_score(top_pick_p["srow"], feat_hdrs, fn)} }},'
        for fn, wt in weights.items() if fn in crit
    )
    json_ld_str = _build_json_ld(list_name, page_slug, ranked, top_score)
    page_data_content = (
        f'// AUTO-GENERATED — do not edit by hand.\n'
        f'// Run: python Documentation/generate_ranked_list.py "{list_name}"\n'
        f'// Generated: {date_iso}\n//\n'
        f'// WHY_TOP_PICK_BULLETS lives in page.tsx — review manually when top pick changes.\n\n'
        f'export const SPREADSHEET_UPDATED = {date_display!r};\n'
        f'export const TOP_SCORE_IN_GROUP = {top_score};\n'
        f'export const TOP_PICK_RAW_SCORE = {top_pick_p["score"]:.2f};\n'
        f'export const TOP_PICK_PERCENT = {top_pick_pct};\n\n'
        f'export const RANKED_SLUGS: string[] = [\n{ranked_lines}\n];\n\n'
        f'export const SCORE_PERCENT: Record<string, number> = {{\n{pct_lines}\n}};\n\n'
        f'// Add special editorial notes only. Price category tags are auto-derived by BestForCard.\n'
        f'export const RUNNER_NOTES: Record<string, string> = {{}};\n\n'
        f'export const TOP_PICK_CRITERIA_DATA: {{ label: string; weight: string; score: number }}[] = [\n'
        f'{crit_lines}\n];\n\n'
        f'// JSON-LD structured data — rendered as <script type="application/ld+json"> in page.tsx\n'
        f'export const JSON_LD = {json_ld_str!r};\n'
    )
    pd_path = os.path.join(page_dir, "page-data.ts")
    with open(pd_path, "w", encoding="utf-8") as f:
        f.write(page_data_content)
    print(f"  ✓  {os.path.relpath(pd_path, REPO_ROOT)}")

    # ── scoring/scoring-data.ts ───────────────────────────────────────────
    os.makedirs(os.path.join(page_dir, "scoring"), exist_ok=True)
    wt_lines = "\n".join(
        f'  {{ key: {crit[fn]["key"]!r}, label: {crit[fn]["short"]!r},'
        f' weight: "{wt:.0%}", reversed: {"true" if crit[fn]["reversed"] else "false"} }},'
        for fn, wt in weights.items() if fn in crit
    )
    row_lines = []
    for i, (sl, pct, w, p) in enumerate(ranked):
        score_parts = ", ".join(
            f'{crit[fn]["key"]}: {_get_score(p["srow"], feat_hdrs, fn)}'
            for fn in weights if fn in crit
        )
        row_lines.append(
            f"  {{\n    rank: {i+1}, manufacturer: {w['mfr_clean']!r},"
            f" product: {w['prod_clean']!r},\n"
            f"    priceCategory: {w.get('price_cat','')!r}, animal: {w.get('animal_cat','')!r},"
            f" price: {'${:.2f}'.format(p['price'])!r},\n"
            f"    rating: {'{:.1f}'.format(p['rating'])!r},"
            f" reviews: {'{:,}'.format(p['reviews'])!r},"
            f" overall: {str(round(w['score'],2))!r},\n"
            f"    scores: {{ {score_parts} }},\n  }},"
        )
    sd_content = (
        f'// AUTO-GENERATED — do not edit by hand.\n'
        f'// Run: python Documentation/generate_ranked_list.py "{list_name}"\n'
        f'// Generated: {date_iso}\n\n'
        f'export const GENERATED_DATE = {date_iso!r};\n\n'
        f'export const WEIGHTS = [\n{wt_lines}\n] as const;\n\n'
        f'export type ScoreKey = typeof WEIGHTS[number]["key"];\n\n'
        f'export type RankedRow = {{\n'
        f'  rank: number; manufacturer: string; product: string;\n'
        f'  priceCategory: string; animal: string; price: string;\n'
        f'  rating: string; reviews: string; overall: string;\n'
        f'  scores: Record<ScoreKey, number>;\n}};\n\n'
        f'export const ROWS: RankedRow[] = [\n' + "\n".join(row_lines) + '\n];\n'
    )
    sd_path = os.path.join(page_dir, "scoring", "scoring-data.ts")
    with open(sd_path, "w", encoding="utf-8") as f:
        f.write(sd_content)
    print(f"  ✓  {os.path.relpath(sd_path, REPO_ROOT)}")

    # ── Template files (new pages only, never overwrite existing) ─────────
    if is_new:
        _write_page_template(page_dir, page_slug, list_name, crit)
        print(f"  → New page created at /{page_slug} — review page.tsx and add content.")


# ── Main ───────────────────────────────────────────────────────────────────
def generate_one(table_name, feat_hdrs, srows, pm, rubric, vis_col, notes_col):
    """
    Runs the full pipeline for a single weighting table and writes its
    output spreadsheet. Returns True on success, False if it failed
    (used by 'all' mode to keep going and report a summary at the end).
    """
    try:
        weights_path, list_name = resolve_weights_path(table_name)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        return False

    out_path = os.path.join(OUT_DIR, f"{list_name} List.xlsx")

    print(f"\n  Weighting table : {weights_path}")
    print(f"  Output          : {out_path}\n")

    weights, filters = load_weights_and_filters(weights_path)

    if filters:
        print(f"Pre-filters (applied before scoring):")
        for field, operator, value in filters:
            print(f"  {field} {operator} {value}")
    print(f"\nWeights ({sum(weights.values()):.0%} total):")
    for fn, wt in weights.items():
        print(f"  {fn}: {wt:.0%}")

    # Build candidates — apply pre-filter first
    candidates = []
    pre_filtered_count = 0
    for (mfr_l, prod_l), srow in srows.items():
        pm_data = pm.get((mfr_l, prod_l), {})
        if filters and not passes_filters(pm_data, filters):
            pre_filtered_count += 1
            continue

        score = calc_score(srow, feat_hdrs, weights)
        vis   = srow.get(vis_col) if vis_col else None
        feat_scores = {}
        for fn in weights:
            for h in feat_hdrs:
                if h and normalise(h) == fn:
                    feat_scores[fn] = srow.get(h)
                    break

        candidates.append({
            "mfr":          srow[feat_hdrs[0]],
            "prod":         srow[feat_hdrs[1]],
            "mfr_clean":    str(srow[feat_hdrs[0]] or "").strip().lstrip(),
            "prod_clean":   str(srow[feat_hdrs[1]] or "").strip(),
            "score":        score,
            "feat_scores":  feat_scores,
            "animal_cat":   pm_data.get("category")  or "Unknown",
            "price_cat":    pm_data.get("price_cat") or "Unknown",
            "rating":       pm_data.get("rating")    or 0,
            "reviews":      pm_data.get("reviews")   or 0,
            "price":        pm_data.get("price"),
            "vis_contrast": vis if isinstance(vis, (int, float)) else 0,
            "notes":        srow.get(notes_col, "") or "",
        })

    if filters:
        print(f"\n  {pre_filtered_count} products removed by pre-filter")
        print(f"  {len(candidates)} products remaining for scoring")

    winners, eliminated = apply_filter(candidates)
    winners.sort(key=lambda x: (x["score"], x["rating"], x["vis_contrast"], x["reviews"]),
                 reverse=True)

    # Console output
    print(f"\n{'='*80}")
    print(f"  {list_name}")
    print(f"  {len(winners)} products in list  |  {len(eliminated)} eliminated by dedup rule")
    print(f"{'='*80}")
    print(f"{'#':<4}{'Score':<7}{'Price Cat':<16}{'Category':<10}{'Rating':<7}{'Manufacturer':<18}Product")
    print("-" * 95)
    for i, w in enumerate(winners, 1):
        print(f"{i:<4}{w['score']:<7.2f}{str(w['price_cat']):<16}{str(w['animal_cat']):<10}"
              f"{str(w['rating']):<7}{w['mfr_clean']:<18}{w['prod_clean']}")

    if eliminated:
        print(f"\n--- Eliminated by dedup rule ({len(eliminated)}) ---")
        for e in sorted(eliminated, key=lambda x: x["score"], reverse=True):
            print(f"  {e['prod_clean']:<44} [{e['animal_cat']}/{e['price_cat']}] "
                  f"score={e['score']:.2f} -> kept: {e.get('eliminated_by','')}")

    write_output(winners, eliminated, pre_filtered_count, weights, filters,
                 rubric, list_name, out_path)
    print(f"\n  Saved: {out_path}")

    # Generate TypeScript data files if this list has a registered page config
    generate_ts_files(list_name, winners, candidates, weights, feat_hdrs, srows, pm)

    return True

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print('ERROR: provide the weighting table name, or "all" to regenerate every table.')
        print('Example: python Documentation\\generate_ranked_list.py "Best for Seniors Living Alone"')
        print('Example: python Documentation\\generate_ranked_list.py all')
        sys.exit(1)

    table_name = " ".join(sys.argv[1:])

    for label, path in [("Product Matrix",        PM_FILE),
                         ("Product Feature Scores", SCORES_FILE),
                         ("Feature Scoring Rubric", RUBRIC_FILE)]:
        if not os.path.exists(path):
            print(f"ERROR: {label} not found:\n  {path}")
            sys.exit(1)

    print(f"  Scores file     : {SCORES_FILE}")
    print(f"  Product Matrix  : {PM_FILE}")

    feat_hdrs, srows = load_scores(SCORES_FILE)
    pm               = load_product_matrix(PM_FILE)
    rubric           = load_rubric(RUBRIC_FILE)
    vis_col          = next((h for h in feat_hdrs if h and "visual" in h.lower()), None)
    notes_col        = next((h for h in feat_hdrs if h and "notes"  in h.lower()), None)

    if table_name.strip().lower() == "all":
        if not os.path.isdir(WT_DIR):
            print(f"ERROR: Best For Weighting Tables folder not found:\n  {WT_DIR}")
            sys.exit(1)

        table_files = sorted(f for f in os.listdir(WT_DIR) if f.endswith(".xlsx") and not f.startswith("~$"))
        if not table_files:
            print(f"ERROR: No .xlsx weighting tables found in:\n  {WT_DIR}")
            sys.exit(1)

        print(f"\n'all' mode — regenerating {len(table_files)} Best For list(s) from:\n  {WT_DIR}\n")

        succeeded, failed = [], []
        for f in table_files:
            name = os.path.splitext(f)[0]
            print(f"\n{'#'*80}\n#  {name}\n{'#'*80}")
            ok = generate_one(name, feat_hdrs, srows, pm, rubric, vis_col, notes_col)
            (succeeded if ok else failed).append(name)

        print(f"\n{'='*80}")
        print(f"  DONE — {len(succeeded)} succeeded, {len(failed)} failed")
        print(f"{'='*80}")
        for name in succeeded:
            print(f"  OK    {name}")
        for name in failed:
            print(f"  FAIL  {name}")

        if failed:
            sys.exit(1)
    else:
        ok = generate_one(table_name, feat_hdrs, srows, pm, rubric, vis_col, notes_col)
        if not ok:
            sys.exit(1)

if __name__ == "__main__":
    main()
